#!/usr/bin/env python3
"""Regenerate everything in outputs/ from the source-of-truth data + reviews.

  python scripts/build_outputs.py

- Word reports (.docx) are produced from reviews/*.md via pandoc.
- Workbooks (.xlsx) are produced from data/*.csv via openpyxl.
- The cross-mechanism gap matrix is recomputed from extraction.csv + axes +
  the species vocabulary (method documented in-sheet).

Binary artefacts are generated, never hand-edited (see methodology/README.md).
No PDFs or copyrighted text are emitted — only derived data and summaries.
"""
import csv, os, re, subprocess
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = os.path.join(ROOT, "data"); REV = os.path.join(ROOT, "reviews"); OUT = os.path.join(ROOT, "outputs")
os.makedirs(OUT, exist_ok=True)
HDR = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="1F7A8C")


def rd(f):
    with open(os.path.join(D, f), encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def sheet_from_rows(ws, rows, headers=None, note=None):
    r0 = 1
    if note:
        ws.cell(1, 1, note).font = Font(italic=True, color="666666"); r0 = 2
    if not rows:
        ws.cell(r0, 1, "(no rows)"); return
    headers = headers or list(rows[0].keys())
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r0, c, h); cell.font = HDR; cell.fill = HFILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, row in enumerate(rows, r0 + 1):
        for c, h in enumerate(headers, 1):
            ws.cell(i, c, row.get(h, "")).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws.cell(r0 + 1, 1)
    ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(headers))}{r0}"
    for c, h in enumerate(headers, 1):
        w = max(len(str(h)), *(len(str(row.get(h, ""))) for row in rows)) if rows else len(h)
        ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 10), 60)


def save(wb, name):
    p = os.path.join(OUT, name); wb.save(p); print("  xlsx:", name)


def matrix_sheet(ws, rowkeys, colkeys, counts, title):
    ws.cell(1, 1, title).font = Font(bold=True)
    ws.cell(2, 1, "").value = ""
    for c, ck in enumerate(colkeys, 2):
        cell = ws.cell(3, c, ck); cell.font = HDR; cell.fill = HFILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(3, 1, "mechanism").font = HDR; ws.cell(3, 1).fill = HFILL
    for r, rk in enumerate(rowkeys, 4):
        ws.cell(r, 1, rk).font = Font(bold=True)
        for c, ck in enumerate(colkeys, 2):
            n = counts.get((rk, ck), 0)
            cell = ws.cell(r, c, n)
            if n == 0:
                cell.fill = PatternFill("solid", fgColor="F4CCCC")  # gap = red
    ws.column_dimensions["A"].width = 26
    for c in range(2, len(colkeys) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 16


# ---------- Word reports via pandoc ----------
DOCX = {
    "state_of_the_art.md": "Fish_passage_injury_mortality_state_of_the_art.docx",
    "barotrauma_overview.md": "Barotrauma_state_of_the_art_overview.docx",
    "collision_overview.md": "Collision_state_of_the_art_overview.docx",
    "shear_overview.md": "Shear_state_of_the_art_overview.docx",
    "cross_mechanism_synthesis.md": "Cross_mechanism_synthesis.docx",
}


def build_docx():
    for md, docx in DOCX.items():
        src = os.path.join(REV, md); dst = os.path.join(OUT, docx)
        subprocess.run(["pandoc", src, "-o", dst, "--toc"], check=True)
        print("  docx:", docx)


# ---------- Workbooks ----------
def build_extraction():
    doi = {r["citation_key"]: r["doi"] for r in rd("corpus.csv")}
    rows = rd("extraction.csv")
    for r in rows:
        r["DOI"] = doi.get(r["citation_key"], "")
    headers = list(rows[0].keys())
    headers.insert(headers.index("Title"), headers.pop(headers.index("DOI")))  # DOI before Title
    wb = Workbook(); sheet_from_rows(wb.active, rows, headers); wb.active.title = "Extraction"
    save(wb, "Fish_passage_injury_mortality_extraction.xlsx")


def build_mechanism(mech, outname):
    wb = Workbook(); first = True
    for label, fn in [("Register", f"{mech}_register.csv"),
                      ("Reproducibility", f"{mech}_reproducibility_scorecard.csv"),
                      ("Metrics", f"{mech}_metrics_catalogue.csv")]:
        if not os.path.exists(os.path.join(D, fn)):
            continue
        ws = wb.active if first else wb.create_sheet(); ws.title = label; first = False
        sheet_from_rows(ws, rd(fn))
    save(wb, outname)


def build_gap_matrix():
    ex = rd("extraction.csv"); ax = rd("axes_exposure_timing.csv"); sp = rd(os.path.join("vocab", "species.csv"))
    # species -> family_group lookup
    fam = {}
    for s in sp:
        for key in (s["common_name"], s["scientific_name"]):
            if key.strip():
                fam[key.strip().lower()] = s["family_group"]
    MECHS = ["Barotrauma/pressure", "Blade strike", "Shear", "Turbulence", "Cavitation",
             "Grinding/abrasion/pinch", "Gas supersaturation/GBT", "Entrainment/impingement/screen"]

    # mechanism x family_group (from extraction)
    fg_counts = defaultdict(int); fgset = set()
    for r in ex:
        mechs = [m.strip() for m in r["Mechanism(s)"].split(";") if m.strip()]
        sptext = r.get("Species", "").lower()
        fams = {fg for name, fg in fam.items() if re.search(r"\b" + re.escape(name) + r"\b", sptext)}
        if not fams:
            fams = {"Unmapped / not stated"}
        for m in mechs:
            for fg in fams:
                fg_counts[(m, fg)] += 1; fgset.add(fg)
    # mechanism (3-class from axes) x environment and x outcome timing
    env_counts = defaultdict(int); envset = set()
    tim_counts = defaultdict(int); timset = set()
    AXM = {"barotrauma": "Barotrauma", "collision": "Collision", "shear": "Shear"}
    for r in ax:
        m = AXM.get(r["mechanisms"].strip())
        if not m:
            continue
        e = r["study_environment"].strip() or "Not stated"; env_counts[(m, e)] += 1; envset.add(e)
        for t in [x.strip() for x in r["outcome_timing"].split(";") if x.strip()]:
            tim_counts[(m, t)] += 1; timset.add(t)

    wb = Workbook()
    matrix_sheet(wb.active, [m for m in MECHS if any(k[0] == m for k in fg_counts)],
                 sorted(fgset), fg_counts,
                 "Studies by mechanism x family group (source: extraction.csv x vocab/species.csv; red = gap)")
    wb.active.title = "Mechanism x family"
    matrix_sheet(wb.create_sheet("Mechanism x environment"), ["Barotrauma", "Collision", "Shear"],
                 sorted(envset), env_counts,
                 "Studies by mechanism x study environment (source: axes_exposure_timing.csv; red = gap)")
    matrix_sheet(wb.create_sheet("Mechanism x outcome timing"), ["Barotrauma", "Collision", "Shear"],
                 sorted(timset), tim_counts,
                 "Studies by mechanism x outcome timing (source: axes_exposure_timing.csv; red = gap)")
    save(wb, "Cross_mechanism_gap_matrix.xlsx")


def main():
    print("Word reports:"); build_docx()
    print("Workbooks:")
    build_extraction()
    build_mechanism("barotrauma", "Barotrauma_metrics_reproducibility.xlsx")
    build_mechanism("collision", "Collision_metrics_reproducibility.xlsx")
    build_mechanism("shear", "Shear_metrics_reproducibility.xlsx")
    build_gap_matrix()
    print("Done.")


if __name__ == "__main__":
    main()
